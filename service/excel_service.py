import csv
import openpyxl
from typing import List, Dict

from repository.job_repository import *

path_to_done = '/bot/storage/done/'
path_to_job_list = '/bot/storage/jobs/'

os.makedirs(path_to_done, exist_ok=True)
os.makedirs(path_to_job_list, exist_ok=True)


def read_xlsx(file_name: str) -> List[List[List[str]]]:
    workbook = openpyxl.load_workbook(file_name)

    return [[list(row) for row in workbook[sheet].values] for sheet in workbook.sheetnames]


def write_xlsx(file_name: str, data: Dict[str, List[List[str]]]):
    workbook = openpyxl.Workbook()

    workbook.remove(workbook['Sheet'])

    for name, rows in data.items():
        sheet = workbook.create_sheet(name)
        for row in rows:
            sheet.append(row)

    workbook.save(file_name)


class ExcelService:
    @staticmethod
    def import_data(data: List[Job]):
        for j in data:
            save_job(j, find_by_params(j.master, j.title) is None)

    @staticmethod
    def import_csv(file_name: str):
        data = []

        def capitalize_first(text: str):
            return text[0].upper() + text[1:]

        with open(path_to_job_list + file_name, 'r') as f:
            r = csv.reader(f, delimiter=',')

            for row in r:
                data.append(Job(-1, row[1], capitalize_first(row[0]), row[2], True, datetime.now()))

        ExcelService.import_data(data)

    @staticmethod
    def rollback(timestamp: datetime):
        ExcelService.import_csv('save_' + str(timestamp) + '.csv')

    @staticmethod
    def delete_all():
        drop_table()

    @staticmethod
    def summarise(start: datetime) -> List[str]:
        total = collect_daily(start)

        res = {}
        for cj in total:
            if cj.job not in res:
                res[cj.job] = 0
            res[cj.job] += cj.count

        return [str(CompletedJob(-1, k, v, datetime.now()))for k, v in res.items()]

    @staticmethod
    def export_csv(start: datetime) -> str:
        return '\n'.join([CompletedJob.csv_title()] + [str(cj) for cj in collect_daily(start)] +\
                         [] + ['?????????? ???? ?????????????? ???????? ??????????'] + ExcelService.summarise(start))

    @staticmethod
    def export_xlsx(start: datetime) -> List[List[str]]:
        return [[e.replace('"', '') for e in CompletedJob.csv_title().split('","')]] +\
               [[e.replace('"', '') for e in str(cj).split('","')] for cj in collect_daily(start)] +\
               [[]] + [['?????????? ???? ?????????????? ???????? ??????????']] +\
               [[e.replace('"', '') for e in str(cj).split('","')] for cj in ExcelService.summarise(start)]

    @staticmethod
    def save(start: datetime, xlsx: bool = False) -> str:
        now = datetime.now()

        file_name = path_to_done + 'save_' + datetime.strftime(now, '%Y%m%d_%H%M%S') + ('.xlsx' if xlsx else '.csv')

        if xlsx:
            write_xlsx(file_name,
                       {f"?????????? ???? {datetime.strftime(now, '%Y-%m-%d (%H-%M)')}": ExcelService.export_xlsx(start)})
        else:
            with open(file_name, 'w', newline='', encoding='utf-8') as f:
                f.write(ExcelService.export_csv(start))

        return file_name
